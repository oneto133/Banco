import openpyxl
import pandas as pd
import shutil
import win32com.client as win32
import sys
import os
from pathlib import Path
from Funções import Tempo, Arquivo

ONEDRIVE = Path(os.environ.get("OneDrive", Path.home()))
BANCO_GENIO = ONEDRIVE / "Banco Gênio"
CSV_DIR = BANCO_GENIO / "csv"

class main(Tempo, Arquivo):
    def __init__(self):
        super().__init__()
        """
        Primeiro se define os arquivos a serem manipulados, após os destinos deverão serem trabalhados
        a origem contém os dados atualizados, há um destino temporário que em tese terá de ser excluído
        ao fim da execução.
        A base é o nosso relatório final.
        """

        # Origem dos dados
        self.origem = BANCO_GENIO / "Banco Gênio.xlsx"
        # Destino temporário
        self.destino = BANCO_GENIO / "Banco_gênio_temp.xlsx"
        # Destino final
        self.base = BANCO_GENIO / "Relatorio.xlsx"
        self.arquivos()

        
    def arquivos(self):
        try:
            df = pd.read_csv(CSV_DIR / "indice.csv" , encoding="latin1")
            shutil.copy2(self.origem, self.destino)
            arq = openpyxl.load_workbook(self.destino, data_only=True)
            base = openpyxl.load_workbook(self.base)
            self.programa(df, arq, base)
        except PermissionError as e:
            print("Parece que o arquivo está aberto, vamos tentar fechar para você")
            try:
                excel = win32.gencache.EnsureDispatch('Excel.Application')
                pasta = excel.Workbooks.Open(self.origem)
                pasta.Close(SaveChanges=False)
                pasta = excel.Workbooks.Open(self.base)
                pasta.Close(SaveChanges=True)
                excel.Quit()
            except Exception as e:
                print("Erro ao tentar fechar a pasta: ", e)


    def programa(self, df, arq, base):

        linha_destino = int(df.columns[0])
        coluna_destino = 1
        Visao_geral = arq["Geral"]
        Banco_de_dados = base["Base"]
        investido = ["B2:B9"]
        atual = ["C2:C9"]
        capital_aplicado = Visao_geral["B11"].value
        capital_real = Visao_geral["C11"].value
        Lucro = Visao_geral["E14"].value
        Capital_total = Visao_geral["C14"].value
        fundo_de_investimento = Visao_geral["C10"].value
        Intervalo = ["A2:A9", "J6:U6"]
        Dados = []
        capital_investido = []
        capital_atual = []
        caixinha_2026 = Visao_geral["C17"].value

        for intervalo in investido:
            for row in Visao_geral[intervalo]:
                valor = [cell.value for cell in row]
                capital_investido.extend(valor)
        
        for intervalo in atual:
            for row in Visao_geral[intervalo]:
                valor = [cell.value for cell in row]
                capital_atual.extend(valor)


        for intervalo in Intervalo:
            for row in Visao_geral[intervalo]:
                dado = [cell.value for cell in row]
                Dados.extend(dado)

        Banco_de_dados.cell(row=linha_destino, column=1, value=capital_aplicado)
        Banco_de_dados.cell(row=linha_destino, column=2, value=Capital_total)
        Banco_de_dados.cell(row=linha_destino, column=3, value=Lucro)
        Banco_de_dados.cell(row=linha_destino, column=5, value=caixinha_2026) #Caixinha
        Banco_de_dados.cell(row=linha_destino, column=6, value=capital_real) #Saldo em caixa
        Banco_de_dados.cell(row=linha_destino, column=19, value=fundo_de_investimento)
        """Preço de mercado, sempre que a quantidade de 'ações' abaixar perto de zero
        aumentar o número de ações, claro que isso vai trazer o preço da ação para baixo, mas
        isso blinda a empresa e atrai mais capital além de limitar a quantidade de investimento em períodos
        o que beneficia para evitar oscilações negativas."""
        Banco_de_dados.cell(row=linha_destino, column=20, value=(Capital_total/8000))

        remap = {4: 6, 5: 8, 6: 9, 7: 10}
        for coluna, valor in enumerate(capital_atual):
            if coluna in remap:
                coluna_final = remap[coluna]
            else:
                coluna_final = coluna
            
            Banco_de_dados.cell(row=linha_destino, column=coluna_destino + (coluna_final + 7), value=valor)

        Banco_de_dados.cell(row=linha_destino, column=coluna+14, value=self.data_atual())
        Banco_de_dados.cell(row=linha_destino, column=coluna+15, value=self.hora_atual())

        with open(CSV_DIR / "indice.csv", "w", encoding="utf-8") as file:
            file.write(f'"{linha_destino + 1}",{self.data_atual()},{self.hora_atual()}')
        base.save(BANCO_GENIO / "Relatorio.xlsx")
        self.excluir(self.destino)

if __name__ == "__main__":
    df = pd.read_csv(CSV_DIR / "indice.csv", encoding="latin1")
    data = df.columns[1]
    Hor = Tempo()
    dat = Hor.data_atual()
    if data != dat:
        main()
    else:
        print("Não executado!")
        res = input("Deseja continuar o programa assim mesmo?")
        if res[0] == "s":   
            main()
